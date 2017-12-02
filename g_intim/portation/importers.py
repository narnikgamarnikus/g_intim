from io import BytesIO
from openpyxl import load_workbook

import requests

from django.conf import settings
from django.core.files import File
from django.core.files.temp import NamedTemporaryFile


from .base import PortationBase
from oscar.core.loading import get_class
from oscar.apps.catalogue.categories import create_from_breadcrumbs

Product = get_class('catalogue.models', 'Product')
Category = get_class('catalogue.models', 'Category')
ProductCategory = get_class('catalogue.models', 'ProductCategory')
ProductAttributeValue = get_class('catalogue.models', 'ProductAttributeValue')
AttributeOption = get_class('catalogue.models', 'AttributeOption')
ProductClass = get_class('catalogue.models', 'ProductClass')
ProductImage = get_class('catalogue.models', 'ProductImage')

Partner = get_class('partner.models', 'Partner')
StockRecord = get_class('partner.models', 'StockRecord')

StockRecord = get_class('catalogue.categories', 'create_from_breadcrumbs')

'''
#from oscar.apps.catalogue.models import Product
from oscar.apps.catalogue.models import Category
from oscar.apps.catalogue.models import ProductCategory
from oscar.apps.catalogue.models import ProductAttributeValue
from oscar.apps.catalogue.models import AttributeOption
from oscar.apps.catalogue.models import ProductClass
from oscar.apps.catalogue.models import ProductImage
from oscar.apps.catalogue.categories import create_from_breadcrumbs

from oscar.apps.catalogue.categories import create_from_breadcrumbs

from oscar.apps.partner.models import Partner
from oscar.apps.partner.models import StockRecord
'''

class CatalogueImporter(PortationBase):

    def __init__(self, file):
        self.wb = load_workbook(BytesIO(file.read()))
        self.ocs = settings.CATEGORIES_SPLIT
        self.ois = settings.IMAGES_SPLIT

    def handle(self):
        self.statistics = {
            'created': 0,
            'updated': 0,
            'errors': [],
        }
        self._import()
        return self.statistics

    def _import(self):
        ws = self.wb.active
        self.max_row = ws.max_row
        for row in ws:
            if row[0].row != 1:
                try:
                    self.create_update_product(row)
                except:
                    self.statistics['errors'].append(str(row[0].row))


    def create_update_product(self, data):

        field_values = data[0:len(self.FIELDS)]
        values = [item.value for item in field_values]
        values = dict(zip(self.FIELDS, values))

        partner = self._get_or_create_partner(values[self.PARTNER])

        product = self._get_or_create_product(values[self.ID], values[self.UPC])
        p_class = self._get_or_create_product_class(values[self.PRODUCT_CLASS])
        
        product = self._product_save(product, p_class, values[self.TITLE], 
                                values[self.DESCRIPTION], values[self.UPC])
        
        if values[self.IMAGE]:
            self._get_or_create_product_image(product, values[self.IMAGE])

        self._save_product_attributes(product, data)

        for category in self._get_or_create_categories(values[self.CATEGORY]):
            product_category = self._product_category_save(product, category)
        
        self._get_or_create_partner_stockrecord(product, partner, values[self.SKU],
                                                values[self.PRICE_RETAIL], 
                                                values[self.COST_PRICE],
                                                values[self.NUM_IN_STOCK]
                                                )

        return product


    def _save_product_attributes(self, product, data):
        
        self.attributes_to_import = product.product_class.attributes.all()
        attrs_values = data[len(self.FIELDS):]
        i = 0
        for attr in self.attributes_to_import:
            try:
                value_obj = product.attribute_values.get(attribute=attr)
            except ProductAttributeValue.DoesNotExist:
                value_obj = ProductAttributeValue()
                value_obj.attribute = attr
                value_obj.product = product
            try:
                value_obj._set_value(attrs_values[i].value)
            except AttributeOption.DoesNotExist:
                attr_option = AttributeOption.objects.create(
                    group=value_obj.attribute.option_group,
                    option=attrs_values[i].value
                )
                value_obj._set_value(attr_option)
            i += 1
            value_obj.save()


    def _get_or_create_categories(self, categories_list):

        categories_list = categories_list.split(self.ocs)

        if isinstance(categories_list, type(None)):
            categories_list = []
        else:
            categories = Category.objects.filter(
            name__in=categories_list)

        if not categories:
            categories = [
                Category.add_root(name=category.strip())
                for category
                in categories_list
            ]

        return categories


    def _get_or_create_product(self, id, upc):
        
        try:
            product = Product.objects.get(id = id)
            self.statistics['updated'] += 1
        except Product.DoesNotExist:
            pass

        try:
            product = Product.objects.get(upc = upc)
            self.statistics['updated'] += 1
        except Product.DoesNotExist:
            product = Product()
            self.statistics['created'] += 1

        return product


    def _get_or_create_product_class(self, name):
        
        product_class, created = ProductClass.objects.get_or_create(name=name)
        
        return product_class


    def _product_save(self, product, product_class, title, description, product_upc):
        
        product.product_class = product_class
        product.title = title
        product.description = description
        product.upc = product_upc

        product.save()

        return product


    def _product_category_save(self, product, category):

        ProductCategory.objects.filter(product=product).delete()

        product_category = ProductCategory()
        
        product_category.product = product
        product_category.category = category
        
        product_category.save()

        return product_category


    def _get_or_create_product_image(self, product, images_list):

            ProductImage.objects.filter(product=product).delete()

            for image in images_list.split(self.ois):
                image = image.strip()
                if image:
                    r = requests.get(image)

                    img_temp = NamedTemporaryFile(delete=True)
                    img_temp.write(r.content)
                    img_temp.flush()

                    image_name = '{}-{}'.format(
                        product.slug,
                        image.split('/')[-1]
                        )

                    product_images = [
                        image.original.name.split('/')[-1] for 
                        image in ProductImage.objects.filter(product=product)
                    ]
                    
                    if not image_name in product_images:
                        self._product_image_save(product, image_name, img_temp)


    def _product_image_save(self, product, image_name, content):

        product_image = ProductImage()
        product_image.product = product
        product_image.original.save(image_name, File(content), save=True)
        product_image.save()
                
        return product_image 


    def _get_or_create_partner(self, parnter_name):
        partner, created = Partner.objects.get_or_create(name=parnter_name)
        return partner

    def _get_or_create_partner_stockrecord(self, product, partner, partner_sku, 
                                            price_retail, cost_price, num_in_stock):
        
        stockrecord, created = StockRecord.objects.get_or_create(product = product,
                                                                partner = partner,
                                                                partner_sku = partner_sku,
                                                                price_retail = price_retail,
                                                                price_excl_tax = price_retail,
                                                                cost_price = cost_price,
                                                                num_in_stock = num_in_stock,
                                                                low_stock_threshold = 5)
        print(stockrecord)
        return stockrecord